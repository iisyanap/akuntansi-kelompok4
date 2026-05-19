import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Konstanta gaya ─────────────────────────────────────────────────────────
FONT_NAME   = "Times New Roman"
FONT_SIZE   = 12
WHITE       = "FFFFFF"
BLACK       = "000000"
LIGHT_GRAY  = "D9D9D9"   # hanya untuk header kolom (abu sangat muda)

def _thin_border():
    s = Side(style="thin", color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thin = Side(style="thin",   color=BLACK)
    thick = Side(style="medium", color=BLACK)
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def _white_fill():
    return PatternFill("solid", fgColor=WHITE)

def _gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def _font(bold=False, size=None, italic=False):
    return Font(name=FONT_NAME, size=size or FONT_SIZE, bold=bold,
                color=BLACK, italic=italic)

def _col(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Judul 3 baris ──────────────────────────────────────────────────────────
def _title_block(ws, nama_perusahaan, nama_bagian, periode, n_cols):
    rows = [
        (nama_perusahaan, True,  14),
        (nama_bagian,     False, 12),
        (periode,         False, 12),
    ]
    for i, (text, bold, size) in enumerate(rows, 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=n_cols)
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT_NAME, size=size, bold=bold, color=BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = size + 6
    # baris kosong pemisah
    ws.row_dimensions[4].height = 6
    return 5   # baris data mulai dari sini

# ── Header kolom ───────────────────────────────────────────────────────────
def _header_row(ws, row_num, headers, col_widths):
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=row_num, column=j, value=h)
        c.font = _font(bold=True)
        c.fill = _gray_fill()
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row_num].height = 20

# ── Sel data biasa ─────────────────────────────────────────────────────────
def _cell(ws, row, col, value, align="left", bold=False,
          num_fmt=None, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, italic=italic)
    c.fill = _white_fill()
    c.alignment = Alignment(horizontal=align, vertical="center",
                            indent=(1 if align=="left" else 0))
    if border:
        c.border = _thin_border()
    if num_fmt:
        c.number_format = num_fmt
    ws.row_dimensions[row].height = 16
    return c

# ── Format angka sesuai mata uang ──────────────────────────────────────────
def _num_fmt(currency):
    if currency == "USD":
        return '"$"#,##0.00'
    return '#,##0'   # Rupiah, tanpa desimal

def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def _num(v):
    if v in ("", None, 0):
        return None
    try:
        return float(str(v).replace(".", "").replace(",", ""))
    except Exception:
        return None

# ══════════════════════════════════════════════════════════════════════════
# 1. JURNAL UMUM
# ══════════════════════════════════════════════════════════════════════════
def export_jurnal(jurnal, perusahaan, currency="IDR"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Umum"
    ws.sheet_view.showGridLines = False

    headers   = ["Tanggal", "Nama Akun", "Ref", "Debit", "Kredit"]
    col_widths = [16, 40, 8, 20, 20]
    n_cols = len(headers)
    nf = _num_fmt(currency)

    r = _title_block(ws, perusahaan["nama"],
                     "Jurnal Umum / General Journal",
                     f"Per {perusahaan['periode']}", n_cols)
    _header_row(ws, r, headers, col_widths); r += 1

    total_d = total_k = 0
    for e in jurnal:
        first = True
        for item in e["debit_entries"]:
            _cell(ws, r, 1, e["tanggal"] if first else "", align="center")
            _cell(ws, r, 2, item["akun"])
            _cell(ws, r, 3, "JU", align="center")
            _cell(ws, r, 4, item["jumlah"], align="right", num_fmt=nf)
            _cell(ws, r, 5, None, align="right")
            total_d += item["jumlah"]; first = False; r += 1
        for item in e["kredit_entries"]:
            _cell(ws, r, 1, "", align="center")
            c = ws.cell(row=r, column=2, value="          " + item["akun"])
            c.font = _font(); c.fill = _white_fill()
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=4)
            ws.row_dimensions[r].height = 16
            _cell(ws, r, 3, "JU", align="center")
            _cell(ws, r, 4, None, align="right")
            _cell(ws, r, 5, item["jumlah"], align="right", num_fmt=nf)
            total_k += item["jumlah"]; r += 1

    # Total
    for j, v in enumerate([None, "T O T A L", None, total_d, total_k], 1):
        c = _cell(ws, r, j, v,
                  align="right" if j > 2 else ("center" if j==1 else "left"),
                  bold=True,
                  num_fmt=nf if j in [4,5] else None)
        c.border = _thick_bottom()

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 2. BUKU BESAR
# ══════════════════════════════════════════════════════════════════════════
def export_buku_besar(buku_besar, perusahaan, currency="IDR"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    nf = _num_fmt(currency)

    headers   = ["Tgl", "Keterangan", "Ref", "Debit", "Kredit", "Saldo D", "Saldo K"]
    col_widths = [14, 32, 8, 18, 18, 18, 18]
    n_cols = len(headers)

    for akun, rows in buku_besar.items():
        ws = wb.create_sheet(title=akun[:31])
        ws.sheet_view.showGridLines = False

        last = rows[-1]
        saldo_d = last.get("saldo_d") or 0
        saldo_k = last.get("saldo_k") or 0
        saldo_info = f"Saldo: {saldo_d or saldo_k:,.0f} ({'D' if saldo_d else 'K'})"

        r = _title_block(ws, perusahaan["nama"],
                         f"Buku Besar — {akun}",
                         f"Per {perusahaan['periode']}  |  {saldo_info}", n_cols)
        _header_row(ws, r, headers, col_widths); r += 1

        for row in rows:
            _cell(ws, r, 1, row["tanggal"], align="center")
            _cell(ws, r, 2, row["keterangan"])
            _cell(ws, r, 3, row.get("source",""), align="center")
            _cell(ws, r, 4, _num(row.get("debit")),   align="right", num_fmt=nf)
            _cell(ws, r, 5, _num(row.get("kredit")),  align="right", num_fmt=nf)
            _cell(ws, r, 6, _num(row.get("saldo_d")), align="right", num_fmt=nf)
            _cell(ws, r, 7, _num(row.get("saldo_k")), align="right", num_fmt=nf)
            r += 1

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 3. NERACA SALDO
# ══════════════════════════════════════════════════════════════════════════
def export_neraca_saldo(neraca_saldo, perusahaan, currency="IDR"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Neraca Saldo"
    ws.sheet_view.showGridLines = False
    nf = _num_fmt(currency)

    headers   = ["No. Rek", "Nama Akun", "Debit", "Kredit"]
    col_widths = [12, 42, 22, 22]
    n_cols = len(headers)

    r = _title_block(ws, perusahaan["nama"],
                     "Neraca Saldo / Trial Balance",
                     f"Per {perusahaan['periode']}", n_cols)
    _header_row(ws, r, headers, col_widths); r += 1

    total_d = total_k = 0
    for row in neraca_saldo:
        _cell(ws, r, 1, row["kode"], align="center")
        _cell(ws, r, 2, row["akun"])
        _cell(ws, r, 3, _num(row["debit"]),   align="right", num_fmt=nf)
        _cell(ws, r, 4, _num(row["kredit"]),  align="right", num_fmt=nf)
        total_d += row["debit"]; total_k += row["kredit"]; r += 1

    for j, v in enumerate([None, "T O T A L", total_d, total_k], 1):
        _cell(ws, r, j, v,
              align="right" if j > 2 else "left", bold=True,
              num_fmt=nf if j > 2 else None).border = _thick_bottom()

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 4. JURNAL PENYESUAIAN
# ══════════════════════════════════════════════════════════════════════════
def export_penyesuaian(jurnal_penyesuaian, perusahaan, currency="IDR"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Penyesuaian"
    ws.sheet_view.showGridLines = False
    nf = _num_fmt(currency)

    headers   = ["Tanggal", "Nama Akun", "Ref", "Debit", "Kredit"]
    col_widths = [16, 40, 8, 20, 20]
    n_cols = len(headers)

    r = _title_block(ws, perusahaan["nama"],
                     "Ayat Jurnal Penyesuaian / Adjusting Journal Entries",
                     f"Per {perusahaan['periode']}", n_cols)
    _header_row(ws, r, headers, col_widths); r += 1

    total_d = total_k = 0
    for e in jurnal_penyesuaian:
        first = True
        for item in e["debit_entries"]:
            _cell(ws, r, 1, e["tanggal"] if first else "", align="center")
            _cell(ws, r, 2, item["akun"])
            _cell(ws, r, 3, "AJP", align="center")
            _cell(ws, r, 4, item["jumlah"], align="right", num_fmt=nf)
            _cell(ws, r, 5, None, align="right")
            total_d += item["jumlah"]; first = False; r += 1
        for item in e["kredit_entries"]:
            _cell(ws, r, 1, "", align="center")
            c = ws.cell(row=r, column=2, value=item["akun"])
            c.font = _font(); c.fill = _white_fill()
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=4)
            ws.row_dimensions[r].height = 16
            _cell(ws, r, 3, "AJP", align="center")
            _cell(ws, r, 4, None, align="right")
            _cell(ws, r, 5, item["jumlah"], align="right", num_fmt=nf)
            total_k += item["jumlah"]; r += 1

    for j, v in enumerate([None, "T O T A L", None, total_d, total_k], 1):
        _cell(ws, r, j, v,
              align="right" if j > 2 else ("center" if j==1 else "left"),
              bold=True,
              num_fmt=nf if j in [4,5] else None).border = _thick_bottom()

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 5. KERTAS KERJA
# ══════════════════════════════════════════════════════════════════════════
def export_kertas_kerja(kertas_kerja, perusahaan, laba_bersih, currency="IDR"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kertas Kerja"
    ws.sheet_view.showGridLines = False
    nf = _num_fmt(currency)

    col_widths = [10, 34, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16]
    n_cols = len(col_widths)

    r = _title_block(ws, perusahaan["nama"],
                     "Kertas Kerja (Worksheet)",
                     f"Per {perusahaan['periode']}", n_cols)

    # Baris grup header
    grp_row = r
    groups = [
        (1,1,"No"), (2,2,"Keterangan"),
        (3,4,"Neraca Saldo"), (5,6,"AJP"),
        (7,8,"NSD"), (9,10,"Laba Rugi"), (11,12,"Neraca"),
    ]
    for c1, c2, label in groups:
        if c1 == c2:
            c = ws.cell(row=grp_row, column=c1, value=label)
        else:
            ws.merge_cells(start_row=grp_row, start_column=c1,
                           end_row=grp_row, end_column=c2)
            c = ws.cell(row=grp_row, column=c1, value=label)
        c.font = _font(bold=True)
        c.fill = _gray_fill()
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[grp_row].height = 18

    # Sub-header D/K
    sub_row = grp_row + 1
    sub_vals = ["","", "D","K","D","K","D","K","D","K","D","K"]
    for j, v in enumerate(sub_vals, 1):
        c = ws.cell(row=sub_row, column=j, value=v)
        c.font = _font(bold=True)
        c.fill = _gray_fill()
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sub_row].height = 16

    # Set lebar kolom
    for j, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    r = sub_row + 1

    num_keys = ["ns_d","ns_k","ajp_d","ajp_k","nsd_d","nsd_k","lr_d","lr_k","ner_d","ner_k"]
    totals = {k: 0 for k in num_keys}

    for row in kertas_kerja:
        _cell(ws, r, 1, row["kode"], align="center")
        _cell(ws, r, 2, row["akun"])
        for k_idx, key in enumerate(num_keys):
            _cell(ws, r, 3+k_idx, _num(row[key]), align="right", num_fmt=nf)
            totals[key] += row[key]
        r += 1

    # Baris laba/rugi
    laba_label = "Laba Bersih" if laba_bersih >= 0 else "Rugi Bersih"
    _cell(ws, r, 1, "", align="center")
    _cell(ws, r, 2, laba_label, bold=True)
    vals_lr = [None,None, None,None, None,None,
               laba_bersih if laba_bersih>=0 else None,
               None if laba_bersih>=0 else abs(laba_bersih),
               None if laba_bersih>=0 else abs(laba_bersih),
               laba_bersih if laba_bersih>=0 else None]
    for k_idx, v in enumerate(vals_lr):
        _cell(ws, r, 3+k_idx, v, align="right", num_fmt=nf, bold=True)
    r += 1

    # Total
    total_vals = [None, "T O T A L",
                  totals["ns_d"], totals["ns_k"],
                  totals["ajp_d"], totals["ajp_k"],
                  totals["nsd_d"], totals["nsd_k"],
                  totals["lr_d"]+(laba_bersih if laba_bersih>0 else 0), totals["lr_k"],
                  totals["ner_d"], totals["ner_k"]+(laba_bersih if laba_bersih>0 else 0)]
    for j, v in enumerate(total_vals, 1):
        c = _cell(ws, r, j, v,
                  align="right" if j>2 else ("center" if j==1 else "left"),
                  bold=True, num_fmt=nf if j>2 else None)
        c.border = _thick_bottom()

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 6. LAPORAN KEUANGAN (3 sheet)
# ══════════════════════════════════════════════════════════════════════════
def export_laporan(laporan, currency="IDR"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    nf = _num_fmt(currency)
    p  = laporan

    # ── Laba Rugi ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Laba Rugi")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24
    lr = p["laba_rugi"]

    r = _title_block(ws, p["nama_perusahaan"],
                     "Laporan Laba Rugi / Income Statement",
                     f"Periode yang Berakhir {p['periode']}", n_cols=2)

    def lr_row(label, value=None, bold=False, indent=0):
        nonlocal r
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _font(bold=bold)
        c1.fill = _white_fill(); c1.border = _thin_border()
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = _font(bold=bold)
        c2.fill = _white_fill(); c2.border = _thin_border()
        c2.number_format = nf
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 16; r += 1

    lr_row("PENDAPATAN", bold=True)
    for akun, jml in lr["pendapatan"]:
        lr_row(akun, jml, indent=2)
    lr_row("Total Pendapatan", lr["total_pendapatan"], bold=True)
    lr_row("")
    lr_row("BEBAN USAHA", bold=True)
    for akun, jml in lr["beban"]:
        lr_row(akun, jml, indent=2)
    lr_row("Total Beban", lr["total_beban"], bold=True)
    lr_row("")
    lbl = "LABA BERSIH" if lr["laba_bersih"]>=0 else "RUGI BERSIH"
    lr_row(lbl, abs(lr["laba_bersih"]), bold=True)

    # ── Perubahan Modal ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Perubahan Modal")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 24
    pm = p["perubahan_modal"]

    r = _title_block(ws2, p["nama_perusahaan"],
                     "Laporan Perubahan Modal",
                     p["periode"], n_cols=2)

    def pm_row(label, value=None, bold=False, indent=0):
        nonlocal r
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.font = _font(bold=bold); c1.fill = _white_fill(); c1.border = _thin_border()
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        c2 = ws2.cell(row=r, column=2, value=value)
        c2.font = _font(bold=bold); c2.fill = _white_fill(); c2.border = _thin_border()
        c2.number_format = nf
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws2.row_dimensions[r].height = 16; r += 1

    pm_row(f"Modal Awal {p['nama_pemilik']}", pm["modal_awal"])
    lbl2 = "(+) Laba Bersih" if pm["laba_bersih"]>=0 else "(-) Rugi Bersih"
    pm_row(lbl2, abs(pm["laba_bersih"]), indent=2)
    if pm["prive"] > 0:
        pm_row(f"(-) Prive {p['nama_pemilik']}", pm["prive"], indent=2)
    pm_row("Modal Akhir", pm["modal_akhir"], bold=True)

    # ── Neraca ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Neraca")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 36
    ws3.column_dimensions["E"].width = 22
    ner = p["neraca"]

    r = _title_block(ws3, p["nama_perusahaan"],
                     "Neraca / Balance Sheet",
                     f"Per {p['periode']}", n_cols=5)

    # Header kiri-kanan
    for j, label in enumerate(["HARTA (ASSETS)","","","KEWAJIBAN & MODAL",""], 1):
        c = ws3.cell(row=r, column=j, value=label)
        c.font = _font(bold=True); c.fill = _gray_fill()
        c.border = _thin_border() if j != 3 else Border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[r].height = 18; r += 1

    total_a = sum(j for _,j in ner["assets"]) - sum(j for _,j in ner["contra_assets"])
    total_l = sum(j for _,j in ner["liabilities"])

    # Isi baris neraca kiri + kanan sejajar
    left_rows  = list(ner["assets"]) + [(f"({a})", -v) for a,v in ner["contra_assets"]] \
                 + [("Total Harta", total_a, True)]
    right_rows = list(ner["liabilities"]) \
                 + [(f"Modal {p['nama_pemilik']}", ner["modal_akhir"])] \
                 + [("Total Kewajiban + Modal", total_l + ner["modal_akhir"], True)]

    max_rows = max(len(left_rows), len(right_rows))
    for i in range(max_rows):
        # kiri
        if i < len(left_rows):
            row_l = left_rows[i]
            bold_l = len(row_l) > 2 and row_l[2]
            indent_l = 0 if bold_l else 2
            c = ws3.cell(row=r, column=1, value=row_l[0])
            c.font = _font(bold=bold_l); c.fill = _white_fill(); c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=indent_l)
            c2 = ws3.cell(row=r, column=2, value=row_l[1] if row_l[1] else None)
            c2.font = _font(bold=bold_l); c2.fill = _white_fill(); c2.border = _thin_border()
            c2.number_format = nf
            c2.alignment = Alignment(horizontal="right", vertical="center")
        else:
            for col in [1,2]:
                c = ws3.cell(row=r, column=col, value="")
                c.fill = _white_fill(); c.border = _thin_border()

        # kolom pemisah
        ws3.cell(row=r, column=3).border = Border()

        # kanan
        if i < len(right_rows):
            row_r = right_rows[i]
            bold_r = len(row_r) > 2 and row_r[2]
            indent_r = 0 if bold_r else 2
            c = ws3.cell(row=r, column=4, value=row_r[0])
            c.font = _font(bold=bold_r); c.fill = _white_fill(); c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=indent_r)
            c2 = ws3.cell(row=r, column=5, value=row_r[1] if row_r[1] else None)
            c2.font = _font(bold=bold_r); c2.fill = _white_fill(); c2.border = _thin_border()
            c2.number_format = nf
            c2.alignment = Alignment(horizontal="right", vertical="center")
        else:
            for col in [4,5]:
                c = ws3.cell(row=r, column=col, value="")
                c.fill = _white_fill(); c.border = _thin_border()

        ws3.row_dimensions[r].height = 16; r += 1

    return _to_bytes(wb)

# ══════════════════════════════════════════════════════════════════════════
# 7. JURNAL PENUTUP
# ══════════════════════════════════════════════════════════════════════════
def export_jurnal_penutup(jurnal_penutup, perusahaan, currency="IDR"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Penutup"
    ws.sheet_view.showGridLines = False
    nf = _num_fmt(currency)

    headers   = ["Tanggal", "Nama Akun", "Ref", "Debit", "Kredit"]
    col_widths = [16, 42, 8, 20, 20]
    n_cols = len(headers)

    r = _title_block(ws, perusahaan["nama"],
                     "Jurnal Penutup / Closing Entries",
                     perusahaan["periode"], n_cols)
    _header_row(ws, r, headers, col_widths); r += 1

    total_d = total_k = 0
    for idx, e in enumerate(jurnal_penutup):
        ref = f"JP{idx+1}"
        first = True
        for item in e["debit_entries"]:
            _cell(ws, r, 1, e["tanggal"] if first else "", align="center")
            _cell(ws, r, 2, item["akun"])
            _cell(ws, r, 3, ref, align="center")
            _cell(ws, r, 4, item["jumlah"], align="right", num_fmt=nf)
            _cell(ws, r, 5, None, align="right")
            total_d += item["jumlah"]; first = False; r += 1
        for item in e["kredit_entries"]:
            _cell(ws, r, 1, "", align="center")
            c = ws.cell(row=r, column=2, value=item["akun"])
            c.font = _font(); c.fill = _white_fill()
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", indent=4)
            ws.row_dimensions[r].height = 16
            _cell(ws, r, 3, ref, align="center")
            _cell(ws, r, 4, None, align="right")
            _cell(ws, r, 5, item["jumlah"], align="right", num_fmt=nf)
            total_k += item["jumlah"]; r += 1

    for j, v in enumerate([None, "T O T A L", None, total_d, total_k], 1):
        c = _cell(ws, r, j, v,
                  align="right" if j>2 else ("center" if j==1 else "left"),
                  bold=True, num_fmt=nf if j in [4,5] else None)
        c.border = _thick_bottom()

    return _to_bytes(wb)
